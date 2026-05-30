"""
Vercel Serverless Function — Export Excel MO 3330
Riceve JSON scenari, compila il template HERM e restituisce il file Excel.
"""
import json, shutil, re, copy, os, io
from pathlib import Path

TEMPLATE = Path(__file__).parent.parent / "TABELLA_RISCHIO_SECONDO_MODELLO_HERM__MO_3330_.xlsx"

AREA_NAMES = {
    'A': 'Requisiti Generali',
    'B': 'Personale',
    'C': 'Acquisti',
    'D': 'Immobilizzazioni',
    'E': 'Contabilità e Reporting Finanziario'
}

CAT_L1_MAP = {
    "Rischi clinico-sanitari": "Rischio Clinico Sanitario",
    "Rischi esterni": "Rischio Esterno",
    "Rischi finanziari": "Rischio Finanziario",
    "Rischi strategici": "Rischio Strategico",
    "Rischi operativi": "Rischio Operativo",
    "Rischi compliance": "Rischio di Compliance",
    "Rischio Finanziario": "Rischio Finanziario",
    "Rischio di Compliance": "Rischio di Compliance",
    "Rischio Operativo": "Rischio Operativo",
    "Rischio Strategico": "Rischio Strategico",
    "Rischio Esterno": "Rischio Esterno",
    "Rischio Clinico Sanitario": "Rischio Clinico Sanitario",
}


def normalizza_cat_l1(v):
    return CAT_L1_MAP.get(v, v or '')


def formatta_cause(c):
    if isinstance(c, dict):
        parti = []
        if c.get("processi"):
            parti.append(f"Processi:\n{c['processi']}")
        if c.get("persone"):
            parti.append(f"Persone:\n{c['persone']}")
        if c.get("fattori_esterni"):
            parti.append(f"Fattori esterni:\n{c['fattori_esterni']}")
        if c.get("strumenti"):
            parti.append(f"Strumenti:\n{c['strumenti']}")
        return "\n\n".join(parti) if parti else ""
    return str(c) if c else ""


def formatta_conseguenze(c):
    if isinstance(c, dict):
        parti = []
        if c.get("economiche"):
            parti.append(f"Economiche: {c['economiche']}")
        if c.get("reputazionali"):
            parti.append(f"Danno d'immagine: {c['reputazionali']}")
        if c.get("compliance"):
            parti.append(f"Compliance normativa: {c['compliance']}")
        if c.get("salute_sicurezza"):
            parti.append(f"Salute e sicurezza: {c['salute_sicurezza']}")
        else:
            parti.append("Salute e sicurezza: -")
        if c.get("gestionale_operativo"):
            parti.append(f"Gestionale - Operativo: {c['gestionale_operativo']}")
        if c.get("ambiente"):
            parti.append(f"Ambiente: {c['ambiente']}")
        return "\n\n".join(parti) if parti else ""
    return str(c) if c else ""


def sv(ws, r, c, v):
    if v is not None and v != "" and v != "null":
        ws.cell(row=r, column=c).value = v


def copy_style(src, tgt):
    if src.has_style:
        tgt.font = copy.copy(src.font)
        tgt.fill = copy.copy(src.fill)
        tgt.border = copy.copy(src.border)
        tgt.alignment = copy.copy(src.alignment)
        tgt.number_format = src.number_format


def clone_row(ws, tmpl_row, new_row):
    offset = new_row - tmpl_row
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=tmpl_row, column=col)
        tgt = ws.cell(row=new_row, column=col)
        copy_style(src, tgt)
        if src.value is None:
            tgt.value = None
        elif isinstance(src.value, str) and src.value.startswith("="):
            formula = re.sub(
                r'(?<!\$)([A-Z]+)(\d+)',
                lambda m: f"{m.group(1)}{int(m.group(2)) + offset}",
                src.value
            )
            tgt.value = formula
        else:
            tgt.value = src.value
    rd = ws.row_dimensions.get(tmpl_row)
    if rd and rd.height:
        ws.row_dimensions[new_row].height = rd.height


def set_wrap(ws, r, cols):
    from openpyxl.styles import Alignment
    for col in cols:
        cell = ws.cell(row=r, column=col)
        al = cell.alignment if cell.alignment else Alignment()
        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal=al.horizontal)


def handler(request):
    """Vercel serverless handler"""
    if request.method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type',
            },
            'body': ''
        }

    if request.method != 'POST':
        return {'statusCode': 405, 'body': 'Method not allowed'}

    try:
        body = json.loads(request.body)
        scenari = body.get('risks', [])
        pac_area = body.get('pacArea', 'D')

        if not scenari:
            return {'statusCode': 400, 'body': json.dumps({"error": "Nessun rischio"})}

        from openpyxl import load_workbook

        # Load template
        shutil.copy2(TEMPLATE, '/tmp/output.xlsx')
        wb = load_workbook('/tmp/output.xlsx')

        area_title = f"PAC AREA {pac_area} {AREA_NAMES.get(pac_area, '').upper()}"

        # Update Copertina title
        if "Copertina" in wb.sheetnames:
            ws_cop = wb["Copertina"]
            for r in range(1, ws_cop.max_row + 1):
                for c in range(1, ws_cop.max_column + 1):
                    v = ws_cop.cell(row=r, column=c).value
                    if v and isinstance(v, str) and "RISK REGISTER" in v:
                        ws_cop.cell(row=r, column=c).value = f"RISK REGISTER - {area_title}"

        # Update Registro dei Rischi title
        ws = wb["Registro dei Rischi"]
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=2, column=c).value
            if v and isinstance(v, str) and "RISK REGISTER" in v:
                ws.cell(row=2, column=c).value = f"RISK REGISTER - {area_title}"

        FIRST_DATA_ROW = 7
        n = len(scenari)

        # Clone rows if needed
        righe_esistenti = ws.max_row - FIRST_DATA_ROW + 1
        if n > righe_esistenti:
            for i in range(righe_esistenti, n):
                clone_row(ws, FIRST_DATA_ROW, FIRST_DATA_ROW + i)

        for idx, s in enumerate(scenari):
            r = FIRST_DATA_ROW + idx

            sv(ws, r, 3, idx + 1)
            sv(ws, r, 5, s.get("direzione", ""))
            sv(ws, r, 6, s.get("risk_owner", ""))
            sv(ws, r, 7, s.get("processo", ""))
            sv(ws, r, 8, normalizza_cat_l1(s.get("categoria_l1", "")))
            sv(ws, r, 9, s.get("categoria_l2", ""))
            sv(ws, r, 10, s.get("categoria_l3", "N.A.") or "N.A.")
            sv(ws, r, 11, s.get("scenario", ""))
            sv(ws, r, 12, formatta_cause(s.get("cause", {})))
            sv(ws, r, 13, formatta_conseguenze(s.get("conseguenze", {})))
            sv(ws, r, 14, s.get("descrizione", ""))
            sv(ws, r, 15, s.get("impatto_economico"))
            sv(ws, r, 16, s.get("impatto_reputazionale"))
            sv(ws, r, 17, s.get("impatto_salute_sicurezza"))
            sv(ws, r, 18, s.get("impatto_compliance"))
            sv(ws, r, 19, s.get("impatto_gestionale_operativo"))
            imp_amb = s.get("impatto_ambiente")
            if imp_amb is not None and imp_amb != "":
                sv(ws, r, 20, imp_amb)
            sv(ws, r, 22, s.get("probabilita_inerente"))
            sv(ws, r, 24, s.get("controlli_correttivi", ""))
            val_imp = s.get("valutazione_controlli_impatto", "")
            for col in [25, 26, 27, 28, 29, 30]:
                sv(ws, r, col, val_imp)
            sv(ws, r, 37, s.get("controlli_preventivi", ""))
            sv(ws, r, 38, s.get("valutazione_controlli_probabilita", ""))
            sv(ws, r, 49, s.get("azioni_mitigazione", ""))
            conf = s.get("confidence_score", 0.5)
            eff = 2 if conf >= 0.7 else (1 if conf >= 0.4 else 0)
            sv(ws, r, 50, eff)
            sv(ws, r, 52, eff)
            set_wrap(ws, r, [11, 12, 13, 14, 24, 37, 49])

        wb.save('/tmp/output.xlsx')

        # Read file and return as base64
        import base64
        with open('/tmp/output.xlsx', 'rb') as f:
            file_data = f.read()

        b64 = base64.b64encode(file_data).decode('utf-8')

        area_file_names = {
            'A': 'Requisiti_Generali', 'B': 'Personale', 'C': 'Acquisti',
            'D': 'Immobilizzazioni', 'E': 'Contabilita_e_Reporting'
        }
        filename = f"Risk_Register_PAC_Area_{pac_area}_{area_file_names.get(pac_area, 'Unknown')}.xlsx"

        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Access-Control-Allow-Origin': '*',
            },
            'body': b64,
            'encoding': 'base64'
        }

    except Exception as e:
        return {
            'statusCode': 500,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({"error": str(e)})
        }
