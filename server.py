"""
HERM Compiler — Server locale
Servono i file statici + gestisce l'export Excel via openpyxl
preservando formule, formattazione e struttura del template originale.
"""
import json, shutil, re, copy, os, sys
from pathlib import Path
from http.server import HTTPServer, SimpleHTTPRequestHandler
from urllib.parse import urlparse, parse_qs

TEMPLATE = Path(__file__).parent / "TABELLA_RISCHIO_SECONDO_MODELLO_HERM__MO_3330_.xlsx"
EXPORT_DIR = Path(__file__).parent / "exports"
EXPORT_DIR.mkdir(exist_ok=True)

AREA_NAMES = {
    'A': 'Requisiti Generali',
    'B': 'Personale',
    'C': 'Acquisti',
    'D': 'Immobilizzazioni',
    'E': 'Contabilità e Reporting Finanziario'
}

CAT_L1_MAP = {
    "Rischi clinico-sanitari": "Rischio Clinico Sanitario",
    "Rischi esterni": "Rischio Esterno",
    "Rischi finanziari": "Rischio Finanziario",
    "Rischi strategici": "Rischio Strategico",
    "Rischi operativi": "Rischio Operativo",
    "Rischi compliance": "Rischio di Compliance",
    "Rischio Finanziario": "Rischio Finanziario",
    "Rischio di Compliance": "Rischio di Compliance",
    "Rischio Operativo": "Rischio Operativo",
    "Rischio Strategico": "Rischio Strategico",
    "Rischio Esterno": "Rischio Esterno",
    "Rischio Clinico Sanitario": "Rischio Clinico Sanitario",
}


def normalizza_cat_l1(v):
    return CAT_L1_MAP.get(v, v or '')


def formatta_cause(c):
    if isinstance(c, dict):
        parti = []
        if c.get("processi"):
            parti.append(f"Processi:\n{c['processi']}")
        if c.get("persone"):
            parti.append(f"Persone:\n{c['persone']}")
        if c.get("fattori_esterni"):
            parti.append(f"Fattori esterni:\n{c['fattori_esterni']}")
        if c.get("strumenti"):
            parti.append(f"Strumenti:\n{c['strumenti']}")
        return "\n\n".join(parti) if parti else ""
    return str(c) if c else ""


def formatta_conseguenze(c):
    if isinstance(c, dict):
        parti = []
        if c.get("economiche"):
            parti.append(f"Economiche: {c['economiche']}")
        if c.get("reputazionali"):
            parti.append(f"Danno d'immagine: {c['reputazionali']}")
        if c.get("compliance"):
            parti.append(f"Compliance normativa: {c['compliance']}")
        if c.get("salute_sicurezza"):
            parti.append(f"Salute e sicurezza: {c['salute_sicurezza']}")
        else:
            parti.append("Salute e sicurezza: -")
        if c.get("gestionale_operativo"):
            parti.append(f"Gestionale - Operativo: {c['gestionale_operativo']}")
        if c.get("ambiente"):
            parti.append(f"Ambiente: {c['ambiente']}")
        return "\n\n".join(parti) if parti else ""
    return str(c) if c else ""


def sv(ws, r, c, v):
    if v is not None and v != "" and v != "null":
        ws.cell(row=r, column=c).value = v


def copy_style(src, tgt):
    if src.has_style:
        tgt.font = copy.copy(src.font)
        tgt.fill = copy.copy(src.fill)
        tgt.border = copy.copy(src.border)
        tgt.alignment = copy.copy(src.alignment)
        tgt.number_format = src.number_format


def clone_row(ws, tmpl_row, new_row):
    offset = new_row - tmpl_row
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=tmpl_row, column=col)
        tgt = ws.cell(row=new_row, column=col)
        copy_style(src, tgt)
        if src.value is None:
            tgt.value = None
        elif isinstance(src.value, str) and src.value.startswith("="):
            formula = re.sub(
                r'(?<!\$)([A-Z]+)(\d+)',
                lambda m: f"{m.group(1)}{int(m.group(2)) + offset}",
                src.value
            )
            tgt.value = formula
        else:
            tgt.value = src.value
    rd = ws.row_dimensions.get(tmpl_row)
    if rd and rd.height:
        ws.row_dimensions[new_row].height = rd.height


def set_wrap(ws, r, cols):
    from openpyxl.styles import Alignment
    for col in cols:
        cell = ws.cell(row=r, column=col)
        al = cell.alignment if cell.alignment else Alignment()
        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal=al.horizontal)


def generate_excel(scenari, pac_area, output_path):
    if not TEMPLATE.exists():
        raise FileNotFoundError(f"Template non trovato: {TEMPLATE}")

    shutil.copy2(TEMPLATE, output_path)
    from openpyxl import load_workbook
    wb = load_workbook(output_path)

    area_title = f"PAC AREA {pac_area} {AREA_NAMES.get(pac_area, '').upper()}"

    # Update Copertina title
    if "Copertina" in wb.sheetnames:
        ws_cop = wb["Copertina"]
        for r in range(1, ws_cop.max_row + 1):
            for c in range(1, ws_cop.max_column + 1):
                v = ws_cop.cell(row=r, column=c).value
                if v and isinstance(v, str) and "RISK REGISTER" in v:
                    ws_cop.cell(row=r, column=c).value = f"RISK REGISTER - {area_title}"

    # Update Registro dei Rischi
    ws = wb["Registro dei Rischi"]
    # Update title in row 2
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v and isinstance(v, str) and "RISK REGISTER" in v:
            ws.cell(row=2, column=c).value = f"RISK REGISTER - {area_title}"

    FIRST_DATA_ROW = 7
    n = len(scenari)

    # Clone rows if needed
    righe_esistenti = ws.max_row - FIRST_DATA_ROW + 1
    if n > righe_esistenti:
        for i in range(righe_esistenti, n):
            clone_row(ws, FIRST_DATA_ROW, FIRST_DATA_ROW + i)

    for idx, s in enumerate(scenari):
        r = FIRST_DATA_ROW + idx

        # Col C (3): # Rischio
        sv(ws, r, 3, idx + 1)
        # Col D (4): ID Rischio → formula CONCAT mantenuta

        # Col E (5): Direzione
        sv(ws, r, 5, s.get("direzione", ""))
        # Col F (6): Risk Owner
        sv(ws, r, 6, s.get("risk_owner", ""))
        # Col G (7): Processo
        sv(ws, r, 7, s.get("processo", ""))
        # Col H (8): Categoria L1
        sv(ws, r, 8, normalizza_cat_l1(s.get("categoria_l1", "")))
        # Col I (9): Categoria L2
        sv(ws, r, 9, s.get("categoria_l2", ""))
        # Col J (10): Categoria L3
        sv(ws, r, 10, s.get("categoria_l3", "N.A.") or "N.A.")
        # Col K (11): Scenario
        sv(ws, r, 11, s.get("scenario", ""))
        # Col L (12): Cause (formatted)
        sv(ws, r, 12, formatta_cause(s.get("cause", {})))
        # Col M (13): Conseguenze (formatted)
        sv(ws, r, 13, formatta_conseguenze(s.get("conseguenze", {})))
        # Col N (14): Descrizione
        sv(ws, r, 14, s.get("descrizione", ""))

        # Impatti inerenti (col 15-20) → O-T
        sv(ws, r, 15, s.get("impatto_economico"))
        sv(ws, r, 16, s.get("impatto_reputazionale"))
        sv(ws, r, 17, s.get("impatto_salute_sicurezza"))
        sv(ws, r, 18, s.get("impatto_compliance"))
        sv(ws, r, 19, s.get("impatto_gestionale_operativo"))
        imp_amb = s.get("impatto_ambiente")
        if imp_amb is not None and imp_amb != "":
            sv(ws, r, 20, imp_amb)
        # Col U (21): Impatto Inerente → FORMULA MAX mantenuta

        # Col V (22): Probabilità Inerente
        sv(ws, r, 22, s.get("probabilita_inerente"))
        # Col W (23): RPN Inerente → FORMULA U*V mantenuta

        # Controlli correttivi (col 24-36) → X-AD
        sv(ws, r, 24, s.get("controlli_correttivi", ""))
        val_imp = s.get("valutazione_controlli_impatto", "")
        for col in [25, 26, 27, 28, 29, 30]:
            sv(ws, r, col, val_imp)
        # Col AE-AJ (31-36): fattori riduzione → VLOOKUP mantenute

        # Controlli preventivi (col 37-39) → AK-AM
        sv(ws, r, 37, s.get("controlli_preventivi", ""))
        sv(ws, r, 38, s.get("valutazione_controlli_probabilita", ""))
        # Col AM (39): fattore riduzione prob → VLOOKUP mantenuta

        # Residuo (col 40-48) → AN-AV → formule mantenute

        # Azioni (col 49-56) → AW-BD
        sv(ws, r, 49, s.get("azioni_mitigazione", ""))
        conf = s.get("confidence_score", 0.5)
        eff = 2 if conf >= 0.7 else (1 if conf >= 0.4 else 0)
        sv(ws, r, 50, eff)   # AX: efficacia azioni correttive
        # Col AY (51): fattore riduzione impatto to-be → VLOOKUP mantenuta
        sv(ws, r, 52, eff)   # AZ: efficacia azioni preventive
        # Col BA (53): fattore riduzione prob to-be → VLOOKUP mantenuta
        # Col BB-BD (54-56): formule to-be mantenute

        set_wrap(ws, r, [11, 12, 13, 14, 24, 37, 49])

    wb.save(output_path)
    return output_path


class HERMHandler(SimpleHTTPRequestHandler):
    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path == '/api/export':
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body)

            scenari = data.get('risks', [])
            pac_area = data.get('pacArea', 'D')

            if not scenari:
                self.send_response(400)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({"error": "Nessun rischio da esportare"}).encode())
                return

            area_names = {
                'A': 'Requisiti_Generali', 'B': 'Personale', 'C': 'Acquisti',
                'D': 'Immobilizzazioni', 'E': 'Contabilità_e_Reporting'
            }
            filename = f"Risk_Register_PAC_Area_{pac_area}_{area_names.get(pac_area, 'Unknown')}.xlsx"
            output_path = EXPORT_DIR / filename

            try:
                generate_excel(scenari, pac_area, output_path)
                with open(output_path, 'rb') as f:
                    file_data = f.read()

                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
                self.send_header('Content-Length', len(file_data))
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(file_data)
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({"error": str(e)}).encode())
            return

        self.send_response(404)
        self.end_headers()

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def end_headers(self):
        if not hasattr(self, '_headers_sent'):
            super().end_headers()


if __name__ == '__main__':
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8080
    os.chdir(Path(__file__).parent)
    server = HTTPServer(('0.0.0.0', port), HERMHandler)
    print(f"HERM Compiler server: http://localhost:{port}")
    print(f"Template: {TEMPLATE}")
    print(f"Export dir: {EXPORT_DIR}")
    server.serve_forever()
