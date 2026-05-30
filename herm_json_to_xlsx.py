"""
HERM JSON → Excel MO 3330
Legge HERM_Risk_Register.json e compila
TABELLA_RISCHIO_SECONDO_MODELLO_HERM__MO_3330_.xlsx
"""
import copy, shutil, re, json, sys, warnings
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment
warnings.filterwarnings("ignore")

TEMPLATE      = Path("TABELLA_RISCHIO_SECONDO_MODELLO_HERM__MO_3330_.xlsx")
JSON_IN       = Path("HERM_Risk_Register.json")
XLSX_OUT      = Path("HERM_MO3330_Compilato.xlsx")
FIRST_DATA_ROW = 7

CAT_L1_MAP = {
    "Rischi clinico-sanitari":   "Rischio Clinico Sanitario",
    "Rischi esterni":            "Rischio Esterno",
    "Rischi finanziari":         "Rischio Finanziario",
    "Rischi strategici":         "Rischio Strategico",
    "Rischi operativi":          "Rischio Operativo",
    "Rischi compliance":         "Rischio di Compliance",
    "Rischio Finanziario":       "Rischio Finanziario",
    "Rischio di Compliance":     "Rischio di Compliance",
    "Rischio Operativo":         "Rischio Operativo",
    "Rischio Strategico":        "Rischio Strategico",
    "Rischio Esterno":           "Rischio Esterno",
    "Rischio Clinico Sanitario": "Rischio Clinico Sanitario",
}

def normalizza_cat_l1(v):
    return CAT_L1_MAP.get(v, v)

def formatta_cause(c):
    if isinstance(c, dict):
        parti = []
        if c.get("processi"):        parti.append(f"Processi:\n{c['processi']}")
        if c.get("persone"):         parti.append(f"Persone:\n{c['persone']}")
        if c.get("fattori_esterni"): parti.append(f"Fattori esterni:\n{c['fattori_esterni']}")
        if c.get("strumenti"):       parti.append(f"Strumenti:\n{c['strumenti']}")
        return "\n\n".join(parti) if parti else ""
    return str(c) if c else ""

def formatta_conseguenze(c):
    if isinstance(c, dict):
        parti = []
        if c.get("economiche"):          parti.append(f"Economiche:\n{c['economiche']}")
        if c.get("reputazionali"):       parti.append(f"Danno d'immagine:\n{c['reputazionali']}")
        if c.get("compliance"):          parti.append(f"Compliance Normativa:\n{c['compliance']}")
        if c.get("salute_sicurezza"):    parti.append(f"Salute e sicurezza:\n{c['salute_sicurezza']}")
        else:                            parti.append("Salute e sicurezza: -")
        if c.get("gestionale_operativo"):parti.append(f"Gestionale - Operativo:\n{c['gestionale_operativo']}")
        if c.get("ambiente"):            parti.append(f"Ambiente:\n{c['ambiente']}")
        return "\n\n".join(parti) if parti else ""
    return str(c) if c else ""

def sv(ws, r, c, v):
    if v is not None and v != "" and v != "null":
        ws.cell(row=r, column=c).value = v

def copy_style(src, tgt):
    if src.has_style:
        tgt.font      = copy.copy(src.font)
        tgt.fill      = copy.copy(src.fill)
        tgt.border    = copy.copy(src.border)
        tgt.alignment = copy.copy(src.alignment)
        tgt.number_format = src.number_format

def clone_row(ws, tmpl_row, new_row):
    offset = new_row - tmpl_row
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=tmpl_row, column=col)
        tgt = ws.cell(row=new_row, column=col)
        copy_style(src, tgt)
        if src.value is None:
            tgt.value = None
        elif isinstance(src.value, str) and src.value.startswith("="):
            formula = re.sub(
                r'(?<!\$)([A-Z]+)(\d+)',
                lambda m: f"{m.group(1)}{int(m.group(2)) + offset}",
                src.value
            )
            tgt.value = formula
        else:
            tgt.value = src.value
    rd = ws.row_dimensions.get(tmpl_row)
    if rd and rd.height:
        ws.row_dimensions[new_row].height = rd.height

def set_wrap(ws, r, cols):
    for col in cols:
        cell = ws.cell(row=r, column=col)
        al = cell.alignment if cell.alignment else Alignment()
        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal=al.horizontal)

def main():
    if not TEMPLATE.exists():
        print(f"ERRORE: template non trovato: {TEMPLATE}"); sys.exit(1)
    if not JSON_IN.exists():
        print(f"ERRORE: JSON non trovato: {JSON_IN}"); sys.exit(1)

    with open(JSON_IN, encoding="utf-8") as f:
        scenari = json.load(f)
    print(f"Letti {len(scenari)} scenari da {JSON_IN}")

    shutil.copy2(TEMPLATE, XLSX_OUT)
    wb = load_workbook(XLSX_OUT)
    ws = wb["Registro dei Rischi"]

    righe_esistenti = ws.max_row - FIRST_DATA_ROW + 1
    n = len(scenari)
    if n > righe_esistenti:
        for i in range(righe_esistenti, n):
            clone_row(ws, FIRST_DATA_ROW, FIRST_DATA_ROW + i)

    for idx, s in enumerate(scenari):
        r = FIRST_DATA_ROW + idx
        ws.cell(row=r, column=3).value = idx + 1
        sv(ws, r, 5,  s.get("direzione", ""))
        sv(ws, r, 6,  s.get("risk_owner", ""))
        sv(ws, r, 7,  s.get("processo", "PAC"))
        sv(ws, r, 8,  normalizza_cat_l1(s.get("categoria_l1", "")))
        sv(ws, r, 9,  s.get("categoria_l2", ""))
        sv(ws, r, 10, s.get("categoria_l3", "N.A.") or "N.A.")
        sv(ws, r, 11, s.get("scenario", ""))
        sv(ws, r, 12, formatta_cause(s.get("cause", "")))
        sv(ws, r, 13, formatta_conseguenze(s.get("conseguenze", "")))
        sv(ws, r, 14, s.get("descrizione", ""))
        sv(ws, r, 15, s.get("impatto_economico"))
        sv(ws, r, 16, s.get("impatto_reputazionale"))
        sv(ws, r, 17, s.get("impatto_salute_sicurezza"))
        sv(ws, r, 18, s.get("impatto_compliance"))
        sv(ws, r, 19, s.get("impatto_gestionale_operativo"))
        sv(ws, r, 20, s.get("impatto_ambiente"))
        sv(ws, r, 22, s.get("probabilita_inerente"))
        sv(ws, r, 24, s.get("controlli_correttivi", ""))
        val_imp = s.get("valutazione_controlli_impatto", "")
        for col in [25, 26, 27, 28, 29, 30]:
            sv(ws, r, col, val_imp)
        sv(ws, r, 37, s.get("controlli_preventivi", ""))
        sv(ws, r, 38, s.get("valutazione_controlli_probabilita", ""))
        sv(ws, r, 49, s.get("azioni_mitigazione", ""))
        conf = s.get("confidence_score", 0.5)
        eff = 2 if conf >= 0.7 else (1 if conf >= 0.4 else 0)
        sv(ws, r, 50, eff)
        sv(ws, r, 52, eff)
        set_wrap(ws, r, [11, 12, 13, 14, 24, 37, 49])

    wb.save(XLSX_OUT)
    print(f"Compilato: {XLSX_OUT} ({n} scenari, fogli: {wb.sheetnames})")

if __name__ == "__main__":
    main()
